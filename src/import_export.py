import openpyxl
from openpyxl.utils import get_column_letter
import duckdb
import os
import shutil
from datetime import datetime
import traceback


class DataManager:
    def __init__(self, db_path):
        self.db_path = db_path

    def _get_connection(self):
        return duckdb.connect(self.db_path)

    def export_to_excel(self, file_path, progress_callback=None):
        """
        Exports all tables (accounts, categories, transactions, budgets) to an Excel file.
        """
        conn = self._get_connection()
        try:
            workbook = openpyxl.Workbook()

            if 'Sheet' in workbook.sheetnames:
                del workbook['Sheet']

            if progress_callback:
                progress_callback(10)

            self._write_table_to_sheet(conn, workbook, 'accounts',
                                       "SELECT * FROM accounts ORDER BY id")
            if progress_callback:
                progress_callback(30)

            self._write_table_to_sheet(conn, workbook, 'categories',
                                       "SELECT * FROM categories ORDER BY category, sub_category")
            if progress_callback:
                progress_callback(50)

            self._write_table_to_sheet(conn, workbook, 'transactions',
                                       "SELECT * FROM transactions ORDER BY date DESC, id DESC")
            if progress_callback:
                progress_callback(80)

            try:
                self._export_exchange_rates_matrix(conn, workbook)
            except Exception as e:
                print(f"Warning: Could not export exchange_rates: {e}")
                traceback.print_exc()

            if progress_callback:
                progress_callback(90)

            try:
                self._write_table_to_sheet(conn, workbook, 'budgets',
                                           "SELECT * FROM budgets ORDER BY category_id")
            except:
                pass

            try:
                self._export_investment_valuations_matrix(conn, workbook)
            except Exception as e:
                print(f"Warning: Could not export investment_valuations: {e}")
                traceback.print_exc()

            workbook.save(file_path)
            if progress_callback:
                progress_callback(100)
            return True, f"Successfully exported data to {file_path}"
        except Exception as e:
            traceback.print_exc()
            return False, f"Export failed: {str(e)}"
        finally:
            conn.close()

    def _write_table_to_sheet(self, conn, workbook, sheet_name, query):
        sheet = workbook.create_sheet(sheet_name)
        result = conn.execute(query).fetchall()

        if not conn.description:
            return

        columns = [col[0] for col in conn.description]

        sheet.append(columns)

        for row in result:
            sheet.append(row)

        for i, col in enumerate(columns):
            sheet.column_dimensions[get_column_letter(i+1)].width = 15

    def import_from_excel(self, file_path, progress_callback=None, skip_backup=False):
        """
        Imports data from an Excel file.
        """
        if not os.path.exists(file_path):
            return False, "File does not exist."

        try:
            if progress_callback:
                progress_callback(5)
            workbook = openpyxl.load_workbook(file_path)
        except Exception as e:
            return False, f"Failed to open Excel file: {str(e)}"

        if progress_callback:
            progress_callback(15)
        valid, msg = self._validate_workbook(workbook)
        if not valid:
            return False, f"Validation failed: {msg}"

        backup_path = None
        if not skip_backup:
            if progress_callback:
                progress_callback(25)
            backup_path = f"{self.db_path}.backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            try:
                shutil.copy2(self.db_path, backup_path)
            except Exception as e:
                return False, f"Failed to create backup: {str(e)}"

        conn = self._get_connection()
        try:
            if progress_callback:
                progress_callback(35)

            conn.execute("DELETE FROM investment_valuations")
            conn.execute("DELETE FROM transactions")
            conn.execute("DELETE FROM budgets")
            conn.execute("DELETE FROM categories")
            conn.execute("DELETE FROM accounts")

            if progress_callback:
                progress_callback(45)

            self._import_sheet(conn, workbook['accounts'], 'accounts')
            if progress_callback:
                progress_callback(55)

            self._import_sheet(conn, workbook['categories'], 'categories')
            if progress_callback:
                progress_callback(65)

            if 'exchange_rates' in workbook.sheetnames:
                conn.execute("DELETE FROM exchange_rates")
                self._import_exchange_rates_matrix(conn, workbook)
            if progress_callback:
                progress_callback(75)

            if 'investment_valuations' in workbook.sheetnames:
                self._import_investment_valuations_matrix(conn, workbook)
            if progress_callback:
                progress_callback(80)

            if 'budgets' in workbook.sheetnames:
                self._import_sheet(conn, workbook['budgets'], 'budgets')
            if progress_callback:
                progress_callback(85)

            self._import_sheet(conn, workbook['transactions'], 'transactions')
            if progress_callback:
                progress_callback(100)

            return True, "Successfully imported data."
        except Exception as e:
            conn.close()
            try:
                if backup_path:
                    shutil.copy2(backup_path, self.db_path)
                return False, f"Import failed: {str(e)}"
            except Exception as restore_error:
                return False, f"Import failed AND restore failed! Error: {str(e)}. Backup at: {backup_path}"
        finally:
            try:
                conn.close()
            except:
                pass

    def _validate_workbook(self, workbook):
        required_sheets = ['accounts', 'categories', 'transactions']
        for sheet in required_sheets:
            if sheet not in workbook.sheetnames:
                return False, f"Missing sheet: {sheet}"

        acc_sheet = workbook['accounts']
        acc_ids = set()

        rows = list(acc_sheet.rows)
        if len(rows) > 0:
            header = [c.value for c in rows[0]]
            if 'id' not in header:
                return False, "Accounts sheet missing 'id' column"
            id_idx = header.index('id')

            for row in rows[1:]:
                val = row[id_idx].value
                if val in acc_ids:
                    return False, f"Duplicate Account ID found: {val}"
                acc_ids.add(val)

        trans_sheet = workbook['transactions']
        rows = list(trans_sheet.rows)
        if len(rows) > 0:
            header = [c.value for c in rows[0]]
            if 'account_id' not in header:
                return False, "Transactions sheet missing 'account_id' column"
            acc_id_idx = header.index('account_id')

            for i, row in enumerate(rows[1:], start=2):
                val = row[acc_id_idx].value
                if val is not None and val not in acc_ids:

                    return False, f"Transaction on row {i} references unknown Account ID: {val}"

        return True, "Valid"

    def _import_sheet(self, conn, sheet, table_name):
        rows = list(sheet.rows)
        if not rows:
            return

        header = [c.value for c in rows[0]]

        placeholders = ', '.join(['?'] * len(header))
        columns = ', '.join(header)
        query = f"INSERT INTO {table_name} ({columns}) VALUES ({placeholders})"

        values_batch = []
        for row in rows[1:]:
            values = [cell.value for cell in row]
            values_batch.append(values)

        if values_batch:
            conn.executemany(query, values_batch)

    def _export_exchange_rates_matrix(self, conn, workbook):
        """
        Exports exchange rates in a matrix format:
        Date | USD | EUR | ...
        """

        data = conn.execute(
            "SELECT date, currency, rate FROM exchange_rates ORDER BY date DESC").fetchall()

        currencies = sorted(list(set(r[1] for r in data)))

        pivot_data = {}
        for row in data:
            date_val = row[0]
            curr = row[1]
            rate = row[2]

            if date_val not in pivot_data:
                pivot_data[date_val] = {}
            pivot_data[date_val][curr] = rate

        sheet = workbook.create_sheet('exchange_rates')
        headers = ['date'] + currencies
        sheet.append(headers)

        sorted_dates = sorted(pivot_data.keys(), reverse=True)

        for d in sorted_dates:
            row = [d]
            for curr in currencies:
                row.append(pivot_data[d].get(curr, None))
            sheet.append(row)

        for col in sheet.columns:
            sheet.column_dimensions[col[0].column_letter].width = 12

    def _import_exchange_rates_matrix(self, conn, workbook):
        sheet = workbook['exchange_rates']
        rows = list(sheet.rows)
        if not rows:
            return

        header = [c.value for c in rows[0]]
        if not header or str(header[0]).lower() != 'date':
            print("Invalid exchange rates header. First column must be 'date'")
            return

        currencies = header[1:]

        full_data = []

        for row in rows[1:]:
            date_val = row[0].value
            if not date_val:
                continue

            if isinstance(date_val, datetime):
                date_str = date_val.strftime('%Y-%m-%d')
            else:
                date_str = str(date_val).split(" ")[0]

            for i, cell in enumerate(row[1:]):
                currency = currencies[i]
                rate = cell.value

                if rate is not None and isinstance(rate, (int, float)) and rate > 0:

                    full_data.append((date_str, currency, float(rate)))

        res = conn.execute(
            "SELECT COALESCE(MAX(id), 0) FROM exchange_rates").fetchone()
        next_id = (res[0] if res else 0) + 1

        final_batch = []
        for item in full_data:
            final_batch.append((next_id, item[0], item[1], item[2]))
            next_id += 1

        conn.executemany(
            "INSERT INTO exchange_rates (id, date, currency, rate) VALUES (?, ?, ?, ?)", final_batch)

    def _export_investment_valuations_matrix(self, conn, workbook):
        """
        Exports investment valuations in keys:
        Date | Account 1 | Account 2 | ...
        Use column names as "Account Name (ID)" to be safe during import map back.
        """

        data = conn.execute("""
            SELECT v.date, a.account, a.id, v.value
            FROM investment_valuations v
            JOIN accounts a ON v.account_id = a.id
            ORDER BY v.date DESC
        """).fetchall()

        accounts_map = {}
        for r in data:
            acc_name = r[1]
            acc_id = r[2]
            token = f"{acc_name}_{acc_id}"
            accounts_map[acc_id] = token

        unique_tokens = sorted(
            list(set(accounts_map.values())), key=lambda x: x.rsplit('_', 1)[0].lower())

        pivot_data = {}
        for r in data:
            date_val = r[0]
            acc_id = r[2]
            val = r[3]
            token = accounts_map[acc_id]

            if date_val not in pivot_data:
                pivot_data[date_val] = {}
            pivot_data[date_val][token] = val

        sheet = workbook.create_sheet('investment_valuations')
        headers = ['date'] + unique_tokens
        sheet.append(headers)

        sorted_dates = sorted(pivot_data.keys(), reverse=True)
        for d in sorted_dates:
            row = [d]
            for token in unique_tokens:
                row.append(pivot_data[d].get(token, None))
            sheet.append(row)

        for col in sheet.columns:
            sheet.column_dimensions[col[0].column_letter].width = 15

    def _import_investment_valuations_matrix(self, conn, workbook):
        sheet = workbook['investment_valuations']
        rows = list(sheet.rows)
        if not rows:
            return

        header = [c.value for c in rows[0]]
        if not header or str(header[0]).lower() != 'date':
            print("Invalid investment_valuations header.")
            return

        token_to_id = {}
        valid_indices = []

        for i, token in enumerate(header[1:], start=1):
            if token and '_' in str(token):
                try:
                    acc_id = int(str(token).rsplit('_', 1)[1])
                    token_to_id[i] = acc_id
                    valid_indices.append(i)
                except:
                    pass

        full_data = []

        for row in rows[1:]:
            date_val = row[0].value
            if not date_val:
                continue

            if isinstance(date_val, datetime):
                date_str = date_val.strftime('%Y-%m-%d')
            else:
                date_str = str(date_val).split(" ")[0]

            for idx in valid_indices:
                val = row[idx].value
                acc_id = token_to_id[idx]

                if val is not None and isinstance(val, (int, float)) and val >= 0:
                    full_data.append((date_str, acc_id, float(val)))

        res = conn.execute(
            "SELECT COALESCE(MAX(id), 0) FROM investment_valuations").fetchone()
        next_id = (res[0] if res else 0) + 1

        final_batch = []
        for item in full_data:
            final_batch.append((next_id, item[0], item[1], item[2]))
            next_id += 1

        conn.executemany(
            "INSERT INTO investment_valuations (id, date, account_id, value) VALUES (?, ?, ?, ?)", final_batch)

    def generate_template(self, file_path):
        """
        Generates a sample Excel template for the user with rich sample data.
        """
        try:
            workbook = openpyxl.Workbook()

            if 'Sheet' in workbook.sheetnames:
                del workbook['Sheet']

            sheet = workbook.create_sheet('accounts')
            sheet.append(['id', 'account', 'type', 'company', 'currency', 'is_active',
                         'show_in_balance', 'is_investment', 'valuation_strategy'])
            sheet.append([0, 'Opening Balance', 'System',
                         None, 'CHF', 1, 0, 0, None])
            sheet.append([1, 'Main Bank Account', 'Bank',
                         'UBS', 'CHF', 1, 1, 0, None])
            sheet.append([2, 'Cash Wallet', 'Cash',
                         None, 'CHF', 1, 1, 0, None])
            sheet.append(
                [3, 'Savings', 'Bank', 'Raiffeisen', 'CHF', 1, 1, 0, None])
            sheet.append([4, 'Global ETF Portfolio', 'Asset',
                         'Interactive Brokers', 'USD', 1, 1, 1, 'Total Value'])
            sheet.append([5, 'McDonalds Stock', 'Asset',
                         'eToro', 'USD', 1, 1, 1, 'Price/Qty'])

            for col in sheet.columns:
                sheet.column_dimensions[col[0].column_letter].width = 20

            sheet = workbook.create_sheet('categories')
            sheet.append(['id', 'sub_category', 'category', 'category_type'])

            sheet.append([1, 'Rent', 'Housing', 'Expense'])
            sheet.append([2, 'Utilities', 'Housing', 'Expense'])
            sheet.append([3, 'Groceries', 'Food', 'Expense'])
            sheet.append([4, 'Restaurants', 'Food', 'Expense'])
            sheet.append([5, 'Public Transport', 'Transport', 'Expense'])
            sheet.append([6, 'Health Insurance', 'Health', 'Expense'])
            sheet.append([7, 'Movies', 'Entertainment', 'Expense'])
            sheet.append([8, 'Salary', 'Income', 'Income'])
            sheet.append([9, 'Dividends', 'Income', 'Income'])

            for col in sheet.columns:
                sheet.column_dimensions[col[0].column_letter].width = 20

            sheet = workbook.create_sheet('transactions')
            headers = ['id', 'date', 'type', 'amount', 'account_id', 'category_id',
                       'payee', 'notes', 'to_account_id', 'to_amount', 'qty', 'invest_account_id']
            sheet.append(headers)

            current_id = 1

            sheet.append([current_id, '2025-01-01', 'transfer', 15000.00, 0, None,
                         'Opening Balance', 'Initial Savings', 1, 15000.00, None, None])
            current_id += 1
            sheet.append([current_id, '2025-01-01', 'transfer', 200.00, 0,
                         None, 'Opening Balance', 'Cash on hand', 2, 200.00, None, None])
            current_id += 1
            sheet.append([current_id, '2025-01-01', 'transfer', 50000.00, 0,
                         None, 'Opening Balance', 'Life Savings', 3, 50000.00, None, None])
            current_id += 1

            sheet.append([current_id, '2025-01-02', 'transfer', 2500.00, 1,
                         None, 'eToro', 'Buy McDonalds (10x)', 5, 2800.00, 10.0, None])
            current_id += 1

            months = ['2025-01', '2025-02', '2025-03']

            for month in months:
                sheet.append([current_id, f'{month}-25', 'income', 6500.00, 1,
                             8, 'Employer Inc.', 'Monthly Salary', None, None, None, None])
                current_id += 1

                sheet.append([current_id, f'{month}-01', 'expense', 1800.00,
                             1, 1, 'Landlord', 'Rent', None, None, None, None])
                current_id += 1
                sheet.append([current_id, f'{month}-02', 'expense', 350.00, 1, 6,
                             'Insurance Co.', 'Health Insurance', None, None, None, None])
                current_id += 1
                sheet.append([current_id, f'{month}-05', 'expense', 80.00,
                             1, 5, 'SBB', 'Monthly Pass', None, None, None, None])
                current_id += 1

                sheet.append([current_id, f'{month}-06', 'transfer', 400.00,
                             1, None, 'ATM', 'Cash Withdrawal', 2, 400.00, None, None])
                current_id += 1

                sheet.append([current_id, f'{month}-08', 'expense', 120.00,
                             1, 3, 'Coop', 'Weekly Groceries', None, None, None, None])
                current_id += 1
                sheet.append([current_id, f'{month}-12', 'expense', 85.00, 1, 4,
                             'Restaurant', 'Dinner with friends', None, None, None, None])
                current_id += 1

                sheet.append([current_id, f'{month}-15', 'expense', 200.00,
                             2, 3, 'Migros', 'Groceries (Cash)', None, None, None, None])
                current_id += 1
                sheet.append([current_id, f'{month}-20', 'expense', 45.00, 2,
                             7, 'Cinema', 'Movie Night (Cash)', None, None, None, None])
                current_id += 1

                sheet.append([current_id, f'{month}-26', 'transfer', 1000.00,
                             1, None, 'Self', 'Monthly Savings', 3, 1000.00, None, None])
                current_id += 1

                sheet.append([current_id, f'{month}-28', 'transfer', 2000.00,
                             1, None, 'IBKR', 'Monthly ETF Buy', 4, 2200.00, 22.5, None])
                current_id += 1

            sheet.append([current_id, '2025-03-15', 'income', 50.00, 4, 9,
                         'Vanguard', 'Quarterly Dividend', None, None, None, None])
            current_id += 1
            sheet.append([current_id, '2025-03-20', 'income', 15.00, 5, 9,
                         'McDonalds', 'Quarterly Dividend', None, None, None, None])
            current_id += 1

            for col in sheet.columns:
                sheet.column_dimensions[col[0].column_letter].width = 15

            sheet = workbook.create_sheet('exchange_rates')
            sheet.append(['date', 'USD', 'EUR'])
            sheet.append(['2025-01-01', 0.88, 0.93])
            sheet.append(['2025-02-01', 0.89, 0.94])
            sheet.append(['2025-03-01', 0.90, 0.95])

            sheet = workbook.create_sheet('budgets')
            sheet.append(['category_id', 'budget_amount'])
            sheet.append([1, 1800.00])
            sheet.append([3, 600.00])
            sheet.append([4, 200.00])
            sheet.append([7, 100.00])

            sheet = workbook.create_sheet('investment_valuations')
            sheet.append(
                ['date', 'Global ETF Portfolio_4', 'McDonalds Stock_5'])
            sheet.append(['2025-01-31', 2250.00, 290.00])
            sheet.append(['2025-02-28', 4600.00, 285.00])
            sheet.append(['2025-03-31', 6800.00, 310.00])

            workbook.save(file_path)
            return True, f"Template created successfully at {file_path}"

        except Exception as e:
            import traceback
            traceback.print_exc()
            return False, f"Failed to create template: {str(e)}"

    def load_sample_data(self):
        """
        Generates and imports sample data into the current database.
        """
        import tempfile

        try:
            fd, path = tempfile.mkstemp(suffix='.xlsx')
            os.close(fd)

            print(f"Generating sample data at {path}...")
            success, msg = self.generate_template(path)
            if not success:
                return False, msg

            print("Importing sample data...")
            def progress(p): pass

            success, msg = self.import_from_excel(
                path, progress, skip_backup=True)

            try:
                os.remove(path)
            except:
                pass

            return success, msg

        except Exception as e:
            return False, str(e)
